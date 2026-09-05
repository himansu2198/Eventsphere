import csv
import re
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from django.utils.timesince import timesince
from django.conf import settings
from groq import Groq
from .decorators import admin_required
from .utils import generate_qr_for_member, notify, notify_admins
from .models import (
    EventCategory, Event, EventUserWish, EventMember, UserMark,
    Profile, EventHistory, Vendor, Notification, Expense, SponsorshipRevenue,
    ContactMessage, ContactMessageReply, Venue, Resource, ResourceAllocation, Sponsor,
    Policy
)


# ---------- Shared helper ----------

def _unread_notif_count(request):
    if request.user.is_authenticated:
        return request.user.notifications.filter(is_read=False).count()
    return 0


def _normalize_search_text(text):
    return re.sub(r'[\s\-]+', '', text or '').lower()


# ---------- Landing ----------

def landing_page(request):
    return render(request, 'landing.html')


# ---------- Auth ----------

def custom_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            profile = getattr(user, 'profile', None)
            if profile and profile.role == 'admin':
                return redirect('dashboard')
            return redirect('participant_dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'registration/login.html')


def participant_signup(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm = request.POST.get('confirm_password', '')
        phone = request.POST.get('phone', '').strip()
        avatar_file = request.FILES.get('avatar')

        errors = []

        if not username:
            errors.append('Username is required.')
        if not email:
            errors.append('Email is required.')
        if not password:
            errors.append('Password is required.')
        if not confirm:
            errors.append('Please confirm your password.')
        if not avatar_file:
            errors.append('Profile photo is required.')

        if password and confirm and password != confirm:
            errors.append('Passwords do not match.')

        if username and User.objects.filter(username=username).exists():
            errors.append('Username already taken.')

        if email and User.objects.filter(email=email).exists():
            errors.append('An account with this email already exists.')

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'participant-signup.html', {
                'form_data': {
                    'username': username,
                    'email': email,
                    'phone': phone,
                },
            })

        user = User.objects.create_user(username=username, email=email, password=password)
        Profile.objects.create(user=user, role='participant', phone=phone, avatar=avatar_file)
        login(request, user)
        messages.success(request, 'Account created successfully!')
        return redirect('participant_dashboard')

    return render(request, 'participant-signup.html')


# ---------- Dashboards ----------

@login_required(login_url='/login/')
def dashboard(request):
    category_count = EventCategory.objects.count()
    event_count = Event.objects.count()
    member_count = EventMember.objects.count()
    unique_participant_count = EventMember.objects.filter(user__isnull=False).values('user').distinct().count()
    completed_count = Event.objects.filter(status='completed').count()
    vendor_count = Vendor.objects.count()
    pending_wishes_count = EventUserWish.objects.filter(status='pending').count()
    total_budget = Event.objects.aggregate(total=Sum('budget'))['total'] or 0
    recent_events = Event.objects.select_related('category').order_by('-id')[:5]

    category_counts = {}
    for ev in Event.objects.select_related('category').all():
        normalized_key = ev.category.name.strip().lower()
        if normalized_key not in category_counts:
            category_counts[normalized_key] = {'label': ev.category.name.strip(), 'count': 0}
        category_counts[normalized_key]['count'] += 1

    sorted_categories = sorted(category_counts.values(), key=lambda c: c['count'], reverse=True)
    chart_labels = [c['label'] for c in sorted_categories]
    chart_data = [c['count'] for c in sorted_categories]

    upcoming_count = Event.objects.filter(status='upcoming').count()
    active_count = Event.objects.filter(status='active').count()

    recent_activities = []

    for h in EventHistory.objects.select_related('event', 'changed_by').order_by('-timestamp')[:10]:
        actor = h.changed_by.username if h.changed_by else 'System'
        action_text = {
            'created': f'{actor} created event "{h.event.name}"',
            'updated': f'{actor} updated event "{h.event.name}"',
            'status_changed': f'{actor} changed status of "{h.event.name}"',
            'cancelled': f'{actor} cancelled "{h.event.name}"',
        }.get(h.action, f'{actor} modified "{h.event.name}"')
        recent_activities.append({
            'text': action_text,
            'timestamp': h.timestamp,
            'icon': 'bi-calendar-event-fill',
            'link': f'/event-detail/{h.event.id}/',
        })

    for m in EventMember.objects.filter(user__isnull=False).select_related('user', 'event').order_by('-joined_at')[:10]:
        recent_activities.append({
            'text': f'{m.user.username} registered for "{m.event.name}"',
            'timestamp': m.joined_at,
            'icon': 'bi-person-plus-fill',
            'link': f'/event-detail/{m.event.id}/',
        })

    for v in Vendor.objects.order_by('-created_at')[:10]:
        recent_activities.append({
            'text': f'Vendor "{v.name}" onboarded',
            'timestamp': v.created_at,
            'icon': 'bi-truck',
            'link': f'/vendors/edit/{v.id}/',
        })

    for e in Expense.objects.filter(status='approved').select_related('event').order_by('-created_at')[:10]:
        recent_activities.append({
            'text': f'Expense of ₹{e.projected_amount} approved for "{e.event.name}"',
            'timestamp': e.created_at,
            'icon': 'bi-cash-coin',
            'link': '/budget-overview/',
        })

    for c in ContactMessage.objects.order_by('-submitted_at')[:10]:
        recent_activities.append({
            'text': f'{c.name} submitted a contact inquiry',
            'timestamp': c.submitted_at,
            'icon': 'bi-envelope-fill',
            'link': f'/contact-messages/{c.id}/',
        })

    recent_activities.sort(key=lambda a: a['timestamp'], reverse=True)
    recent_activities = recent_activities[:8]

    total_actual_spend = Expense.objects.filter(status='approved').aggregate(
        total=Sum('actual_amount'))['total'] or 0
    utilization_pct = round((total_actual_spend / total_budget) * 100, 1) if total_budget else 0

    insights = []

    if pending_wishes_count > 0:
        insights.append({
            'type': 'warning',
            'title': 'Pending Approvals',
            'text': f'{pending_wishes_count} event wish{"es" if pending_wishes_count != 1 else ""} awaiting your review.',
        })
    else:
        insights.append({
            'type': 'success',
            'title': 'All Caught Up',
            'text': 'No pending event wishes right now.',
        })

    insights.append({
        'type': 'info',
        'title': 'Event Status',
        'text': f'{active_count} active, {upcoming_count} upcoming, {completed_count} completed.',
    })

    if total_budget > 0:
        insights.append({
            'type': 'primary',
            'title': 'Budget Utilization',
            'text': f'₹{total_actual_spend} spent of ₹{total_budget} allocated ({utilization_pct}%).',
        })
    else:
        insights.append({
            'type': 'primary',
            'title': 'Budget',
            'text': 'No budget allocated yet.',
        })

    return render(request, 'dashboard.html', {
        'category_count': category_count,
        'event_count': event_count,
        'member_count': member_count,
        'unique_participant_count': unique_participant_count,
        'completed_count': completed_count,
        'vendor_count': vendor_count,
        'pending_wishes_count': pending_wishes_count,
        'total_budget': total_budget,
        'recent_events': recent_events,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'insights': insights,
        'unread_notification_count': _unread_notif_count(request),
        'upcoming_count': upcoming_count,
        'recent_activities': recent_activities,
    })


@login_required(login_url='/login/')
def participant_dashboard(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'admin':
        return redirect('dashboard')

    my_registrations = EventMember.objects.filter(user=request.user).select_related('event')
    registered_event_ids = list(my_registrations.values_list('event_id', flat=True))

    return render(request, 'participant-dashboard.html', {
        'registrations': my_registrations,
        'registered_event_ids': registered_event_ids,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def participant_event_list(request):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'admin':
        return redirect('dashboard')

    today = timezone.now().date()
    events = (
        Event.objects.filter(end_date__gte=today)
        .exclude(status='completed')
        .select_related('category')
        .order_by('start_date')
    )
    registered_event_ids = list(
        EventMember.objects.filter(user=request.user).values_list('event_id', flat=True)
    )

    return render(request, 'participant-event-list.html', {
        'events': events,
        'registered_event_ids': registered_event_ids,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def participant_events_calendar_api(request):
    """
    Returns events in FullCalendar's expected JSON format, pulled fresh
    from the Event model on every request. No caching, no duplicate
    calendar data. Any admin edit to date/time/name is reflected
    immediately since this queries the live database each time.

    IMPORTANT: FullCalendar treats 'end' as EXCLUSIVE, but our
    Event.end_date is stored as the INCLUSIVE last day of the event.
    We add one day to end_date (for all-day events, i.e. no end_time
    set) so the event correctly displays through its actual last day.
    """
    events = Event.objects.exclude(status='completed').select_related('category')
    data = []

    for ev in events:
        if ev.start_time:
            start_value = f"{ev.start_date}T{ev.start_time}"
        else:
            start_value = str(ev.start_date)

        if ev.end_time:
            end_value = f"{ev.end_date}T{ev.end_time}"
        else:
            exclusive_end = ev.end_date + timedelta(days=1)
            end_value = str(exclusive_end)

        data.append({
            'id': ev.id,
            'title': ev.name,
            'start': start_value,
            'end': end_value,
            'url': f'/events-schedule/{ev.id}/',
            'extendedProps': {
                'category': ev.category.name,
                'status': ev.status,
                'venue': ev.venue or 'TBA',
            },
        })

    response = JsonResponse(data, safe=False)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required(login_url='/login/')
def participant_event_detail(request, pk):
    profile = getattr(request.user, 'profile', None)
    if profile and profile.role == 'admin':
        return redirect('dashboard')

    event = get_object_or_404(Event, pk=pk)
    member = EventMember.objects.filter(event=event, user=request.user).first()
    already_registered = member is not None

    attendance_status = None
    if member:
        latest_mark = member.marks.order_by('-marked_at').first()
        attendance_status = latest_mark.status if latest_mark else 'not_marked'

    return render(request, 'participant-event-detail.html', {
        'event': event,
        'already_registered': already_registered,
        'attendance_status': attendance_status,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def participant_register_event(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    profile = getattr(request.user, 'profile', None)

    if not profile or profile.role != 'participant':
        messages.error(request, 'Only participants can register for events.')
        return redirect('dashboard')

    existing = EventMember.objects.filter(event=event, user=request.user).first()
    if existing:
        messages.info(request, 'You are already registered for this event.')
        return redirect('participant_dashboard')

    if event.registration_deadline and timezone.now() > event.registration_deadline:
        messages.error(request, f'Registration for "{event.name}" closed on {event.registration_deadline.strftime("%d %b %Y, %H:%M")}.')
        return redirect('participant_dashboard')

    if event.max_attendees > 0:
        current_count = event.members.count()
        if current_count >= event.max_attendees:
            messages.error(request, f'"{event.name}" has reached its maximum capacity of {event.max_attendees} attendees.')
            return redirect('participant_dashboard')

    member = EventMember.objects.create(
        event=event,
        user=request.user,
        name=request.user.get_full_name() or request.user.username,
        email=request.user.email,
        role='general',
    )
    generate_qr_for_member(member)

    notify_admins(f'{request.user.username} registered for "{event.name}".', link=f'/event-detail/{event.id}/')

    messages.success(request, 'Registered successfully! Your QR ticket is ready.')
    return redirect('participant_dashboard')


@login_required(login_url='/login/')
def participant_cancel_registration(request, member_id):
    member = get_object_or_404(EventMember, pk=member_id, user=request.user)
    if request.method == 'POST':
        event_name = member.event.name
        member.delete()
        messages.success(request, f'Your registration for "{event_name}" has been cancelled.')
        return redirect('participant_dashboard')
    return render(request, 'cancel-registration.html', {'member': member})


@login_required(login_url='/login/')
def participant_search_api(request):
    query = request.GET.get('q', '').strip()
    results = []

    if len(query) >= 2:
        normalized_query = _normalize_search_text(query)
        seen_event_ids = set()

        my_registrations = EventMember.objects.filter(
            user=request.user
        ).select_related('event', 'event__category')

        for reg in my_registrations:
            if reg.event.id in seen_event_ids:
                continue
            if normalized_query in _normalize_search_text(reg.event.name):
                seen_event_ids.add(reg.event.id)
                results.append({
                    'type': 'My Ticket',
                    'label': f'{reg.event.name} ({reg.event.category.name})',
                    'url': f'/events-schedule/{reg.event.id}/',
                })

        other_events = Event.objects.exclude(id__in=seen_event_ids).select_related('category')
        for ev in other_events:
            if normalized_query in _normalize_search_text(ev.name):
                results.append({
                    'type': 'Event',
                    'label': f'{ev.name} ({ev.category.name})',
                    'url': f'/events-schedule/{ev.id}/',
                })

        venues = Venue.objects.all()
        for v in venues:
            if normalized_query in _normalize_search_text(v.name):
                results.append({
                    'type': 'Venue',
                    'label': f'{v.name} ({v.get_venue_type_display()})',
                    'url': f'/events-schedule/',
                })

    return JsonResponse({'results': results[:15]})


@admin_required
def qr_checkin(request):
    result = None
    checked_in_member = None
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        mode = request.POST.get('mode', 'entry')
        event_id = request.POST.get('event_id') or None

        member = EventMember.objects.select_related('event').filter(ticket_code=code).first()

        if not member:
            result = "Invalid ticket code."
        elif event_id and str(member.event_id) != str(event_id):
            result = f"This ticket belongs to \"{member.event.name}\", not the selected event."
        else:
            if mode == 'exit':
                mark = UserMark.objects.filter(member=member, status='present').order_by('-marked_at').first()
                if not mark:
                    result = f"{member.name} has not checked in yet — cannot check out."
                elif mark.checked_out_at:
                    result = f"Already checked out: {member.name} ({member.event.name})"
                else:
                    mark.checked_out_at = timezone.now()
                    mark.save()
                    result = f"Checked out: {member.name} ({member.event.name})"
                    checked_in_member = member
            else:
                if not member.is_eligible_for_checkin:
                    result = (
                        f"🔴 Check-in blocked: {member.name}'s registration status is "
                        f"\"{member.get_registration_status_display()}\", which is not eligible for entry. "
                        f"Only Confirmed registrations can check in."
                    )
                else:
                    already = UserMark.objects.filter(member=member, status='present', checked_out_at__isnull=True).exists()
                    if already:
                        result = "This participant has already been checked in."
                    else:
                        UserMark.objects.create(member=member, status='present')
                        result = f"Checked in: {member.name} ({member.event.name})"
                        checked_in_member = member

        if is_ajax:
            data = {'message': result, 'success': checked_in_member is not None}
            if checked_in_member:
                latest_mark = checked_in_member.marks.order_by('-marked_at').first()
                data['attendee'] = {
                    'name': checked_in_member.name,
                    'email': checked_in_member.email,
                    'phone': checked_in_member.phone or '-',
                    'role': checked_in_member.get_role_display(),
                    'department': checked_in_member.department or '-',
                    'ticket_code': checked_in_member.ticket_code,
                    'event': checked_in_member.event.name,
                    'event_id': checked_in_member.event_id,
                    'registration_status': checked_in_member.get_registration_status_display(),
                    'in_venue': (latest_mark.checked_out_at is None) if latest_mark else False,
                    'time': timezone.localtime(
                        latest_mark.checked_out_at or latest_mark.marked_at
                    ).strftime('%I:%M %p') if latest_mark else '',
                    'was_exit': mode == 'exit',
                }
            return JsonResponse(data)

    selected_event_id = request.GET.get('event')
    checkins_qs = (
        UserMark.objects.filter(status='present')
        .select_related('member', 'member__event')
        .order_by('-marked_at')
    )
    if selected_event_id:
        checkins_qs = checkins_qs.filter(member__event_id=selected_event_id)

    recent_checkins = checkins_qs[:30]

    events_with_counts = (
        Event.objects.filter(status__in=['active', 'upcoming'])
        .annotate(verified_count=Count('members__marks', filter=Q(members__marks__status='present'), distinct=True))
        .order_by('name')
    )

    return render(request, 'qr-checkin.html', {
        'result': result,
        'checked_in_member': checked_in_member,
        'recent_checkins': recent_checkins,
        'events_with_counts': events_with_counts,
        'selected_event_id': int(selected_event_id) if selected_event_id else None,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Global Admin Search (results page) ----------

@admin_required
def global_admin_search(request):
    query = request.GET.get('q', '').strip()

    events = []
    categories = []
    vendors = []
    venues = []
    resources = []
    sponsors = []
    total_results = 0

    if query:
        events = Event.objects.filter(name__icontains=query).select_related('category')[:20]
        categories = EventCategory.objects.filter(name__icontains=query)[:20]
        vendors = Vendor.objects.filter(name__icontains=query)[:20]
        venues = Venue.objects.filter(name__icontains=query)[:20]
        resources = Resource.objects.filter(name__icontains=query)[:20]
        sponsors = Sponsor.objects.filter(name__icontains=query)[:20]

        total_results = (
            events.count() + categories.count() + vendors.count() +
            venues.count() + resources.count() + sponsors.count()
        )

    return render(request, 'admin-search-results.html', {
        'query': query,
        'events': events,
        'categories': categories,
        'vendors': vendors,
        'venues': venues,
        'resources': resources,
        'sponsors': sponsors,
        'total_results': total_results,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Global Search API (live navbar dropdown, admin) ----------

@admin_required
def global_search_api(request):
    query = request.GET.get('q', '').strip()
    results = []

    if len(query) >= 2:
        for ev in Event.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Event', 'label': ev.name, 'url': f'/event-detail/{ev.id}/'})

        for cat in EventCategory.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Category', 'label': cat.name, 'url': f'/category-edit/{cat.id}/'})

        for m in EventMember.objects.filter(name__icontains=query).select_related('event')[:5]:
            results.append({'type': 'Participant', 'label': f'{m.name} — {m.event.name}', 'url': f'/event-detail/{m.event.id}/'})

        for v in Vendor.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Vendor', 'label': v.name, 'url': f'/vendors/edit/{v.id}/'})

        for ven in Venue.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Venue', 'label': ven.name, 'url': f'/venues/edit/{ven.id}/'})

        for r in Resource.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Resource', 'label': r.name, 'url': f'/resources/edit/{r.id}/'})

        for s in Sponsor.objects.filter(name__icontains=query)[:5]:
            results.append({'type': 'Sponsor', 'label': s.name, 'url': f'/sponsors/edit/{s.id}/'})

        for c in ContactMessage.objects.filter(Q(name__icontains=query) | Q(subject__icontains=query))[:5]:
            results.append({'type': 'Contact Inquiry', 'label': f'{c.name} — {c.subject or "No subject"}', 'url': f'/contact-messages/{c.id}/'})

    return JsonResponse({'results': results[:20]})


# ---------- Profile / Settings ----------

@login_required(login_url='/login/')
def profile_view(request):
    profile, created = Profile.objects.get_or_create(user=request.user, defaults={'role': 'participant'})

    if request.method == 'POST':
        if request.FILES.get('avatar'):
            profile.avatar = request.FILES.get('avatar')
            profile.save()
            messages.success(request, 'Profile photo updated!')

        full_name = request.POST.get('full_name', '').strip()
        if full_name:
            name_parts = full_name.split(' ', 1)
            request.user.first_name = name_parts[0]
            request.user.last_name = name_parts[1] if len(name_parts) > 1 else ''

        new_email = request.POST.get('email')
        if new_email and new_email != request.user.email:
            if User.objects.filter(email=new_email).exclude(pk=request.user.pk).exists():
                messages.error(request, 'That email is already in use by another account.')
                return redirect('profile')
            request.user.email = new_email

        request.user.save()

        new_phone = request.POST.get('phone', '').strip()
        profile.phone = new_phone
        profile.save()

        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')

    back_url = 'dashboard' if profile.role == 'admin' else 'participant_dashboard'
    return render(request, 'profile.html', {
        'profile': profile,
        'back_url': back_url,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def settings_view(request):
    return render(request, 'settings.html', {
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def notifications_view(request):
    notes = request.user.notifications.order_by('-created_at')[:50]
    unread_count = request.user.notifications.filter(is_read=False).count()
    return render(request, 'notifications.html', {
        'notifications': notes,
        'unread_count': unread_count,
        'unread_notification_count': unread_count,
    })


@login_required(login_url='/login/')
def mark_notification_read(request, pk):
    note = get_object_or_404(Notification, pk=pk, user=request.user)
    note.is_read = True
    note.save()
    if note.link:
        return redirect(note.link)
    return redirect('notifications')


@login_required(login_url='/login/')
def mark_all_notifications_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    messages.success(request, 'All notifications marked as read.')
    return redirect('notifications')


@login_required(login_url='/login/')
def notification_unread_count_api(request):
    count = request.user.notifications.filter(is_read=False).count()
    return JsonResponse({'unread_count': count})


@login_required(login_url='/login/')
def notification_recent_api(request):
    notes = request.user.notifications.order_by('-created_at')[:8]
    data = []
    for n in notes:
        link = n.link or ''
        if 'contact-messages' in link:
            ntype = 'inquiry'
            title = 'New Contact Inquiry'
        elif 'my-inquiries' in link:
            ntype = 'inquiry'
            title = 'Reply Received'
        elif 'event-detail' in link:
            ntype = 'event'
            title = 'Event Registration'
        else:
            ntype = 'event'
            title = 'Notification'

        data.append({
            'id': n.id,
            'title': title,
            'type': ntype,
            'message': n.message,
            'link': n.link,
            'is_read': n.is_read,
            'timesince': timesince(n.created_at),
        })

    return JsonResponse({
        'notifications': data,
        'unread_count': request.user.notifications.filter(is_read=False).count(),
    })


@login_required(login_url='/login/')
def messages_view(request):
    return render(request, 'messages.html', {
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def apps_menu_view(request):
    return render(request, 'apps-menu.html', {
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def change_password_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Password changed successfully!')
            return redirect('profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'change-password.html', {
        'form': form,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Event Category ----------

@admin_required
def create_category(request):
    if request.method == 'POST':
        EventCategory.objects.create(
            name=request.POST.get('name'),
            code=request.POST.get('code'),
            image=request.FILES.get('image'),
            priority=request.POST.get('priority') or 1,
            status=request.POST.get('status'),
        )
        messages.success(request, 'Category saved successfully!')
        return redirect('category_list')
    return render(request, 'create-category.html', {
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def category_list(request):
    categories = EventCategory.objects.all().order_by('priority')
    return render(request, 'category-list.html', {
        'categories': categories,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_category(request, pk):
    category = get_object_or_404(EventCategory, pk=pk)
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.code = request.POST.get('code')
        if request.FILES.get('image'):
            category.image = request.FILES.get('image')
        category.priority = request.POST.get('priority') or 1
        category.status = request.POST.get('status')
        category.save()
        messages.success(request, 'Category updated successfully!')
        return redirect('category_list')
    return render(request, 'edit-category.html', {
        'category': category,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def category_delete_confirm(request, pk):
    category = get_object_or_404(EventCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted.')
        return redirect('category_list')
    return render(request, 'category-delete.html', {
        'category': category,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Events ----------

def _check_venue_capacity(venue_ref, max_attendees):
    if not venue_ref or not max_attendees:
        return None
    try:
        max_attendees = int(max_attendees)
    except (TypeError, ValueError):
        return None
    if venue_ref.capacity and max_attendees > venue_ref.capacity:
        return f"⚠️ Capacity exceeded. {venue_ref.name} can accommodate only {venue_ref.capacity} attendees."
    return None


def _check_venue_schedule_conflict(venue_ref, start_date, end_date, start_time, end_time, exclude_pk=None):
    if not venue_ref:
        return None

    qs = Event.objects.filter(venue_ref=venue_ref, start_date__lte=end_date, end_date__gte=start_date)
    if exclude_pk:
        qs = qs.exclude(pk=exclude_pk)

    if not start_time or not end_time:
        conflict = qs.first()
        if conflict:
            return f"🔴 Venue unavailable. {venue_ref.name} is already booked for \"{conflict.name}\" on an overlapping date."
        return None

    def _to_date(value):
        if isinstance(value, str):
            return datetime.strptime(value, '%Y-%m-%d').date()
        return value

    def _to_time(value):
        if isinstance(value, str):
            return datetime.strptime(value, '%H:%M').time()
        return value

    try:
        new_start = datetime.combine(_to_date(start_date), _to_time(start_time))
        new_end = datetime.combine(_to_date(end_date), _to_time(end_time))
    except (ValueError, TypeError):
        return None

    for other in qs:
        if not other.start_time or not other.end_time:
            continue
        other_start = datetime.combine(other.start_date, other.start_time)
        other_end = datetime.combine(other.end_date, other.end_time)

        if new_start < other_end and other_start < new_end:
            return (
                f"🔴 Venue unavailable. {venue_ref.name} is already booked for \"{other.name}\" "
                f"from {other.start_time.strftime('%I:%M %p')} – {other.end_time.strftime('%I:%M %p')}."
            )

    return None


@admin_required
def create_event(request):
    categories = EventCategory.objects.all()
    venues = Venue.objects.filter(status='available')
    if request.method == 'POST':
        category_id = request.POST.get('category')
        if not category_id:
            messages.error(request, 'Please select a category.')
            return render(request, 'create-event.html', {
                'categories': categories, 'venues': venues,
                'unread_notification_count': _unread_notif_count(request),
            })
        try:
            venue = request.POST.get('venue')
            venue_ref_id = request.POST.get('venue_ref') or None
            start_date_str = request.POST.get('start_date')
            end_date_str = request.POST.get('end_date')
            start_time = request.POST.get('start_time') or None
            end_time = request.POST.get('end_time') or None
            max_attendees = request.POST.get('max_attendees') or 0

            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if start_date_obj < timezone.now().date():
                messages.error(request, "⚠️ This event's start date is in the past. Please choose today's date or a future date.")
                return render(request, 'create-event.html', {
                    'categories': categories, 'venues': venues,
                    'unread_notification_count': _unread_notif_count(request),
                })

            venue_ref_obj = Venue.objects.filter(pk=venue_ref_id).first() if venue_ref_id else None

            if venue_ref_obj:
                capacity_warning = _check_venue_capacity(venue_ref_obj, max_attendees)
                if capacity_warning:
                    messages.error(request, capacity_warning)
                    return render(request, 'create-event.html', {
                        'categories': categories, 'venues': venues,
                        'unread_notification_count': _unread_notif_count(request),
                    })

                conflict_warning = _check_venue_schedule_conflict(
                    venue_ref_obj, start_date_str, end_date_str, start_time, end_time
                )
                if conflict_warning:
                    messages.error(request, conflict_warning)
                    return render(request, 'create-event.html', {
                        'categories': categories, 'venues': venues,
                        'unread_notification_count': _unread_notif_count(request),
                    })

            event = Event.objects.create(
                name=request.POST.get('name'),
                category_id=category_id,
                description=request.POST.get('description'),
                priority=request.POST.get('priority') or 1,
                scheduled_status=request.POST.get('scheduled_status'),
                venue=venue,
                venue_ref_id=venue_ref_id,
                start_date=start_date_str,
                end_date=end_date_str,
                location=request.POST.get('location'),
                points=request.POST.get('points') or 0,
                max_attendees=max_attendees,
                registration_deadline=request.POST.get('registration_deadline') or None,
                status=request.POST.get('status'),
                image=request.FILES.get('image'),
                session_name=request.POST.get('session_name'),
                speaker_name=request.POST.get('speaker_name'),
                start_time=start_time,
                end_time=end_time,
                venue_name=request.POST.get('venue_name'),
                budget=request.POST.get('budget') or 0,
                sponsors=request.POST.get('sponsors', ''),
            )

            EventHistory.objects.create(event=event, action='created', changed_by=request.user, notes='Event created.')

            messages.success(request, 'Event created successfully!')
            return redirect('event_list')
        except Exception as e:
            messages.error(request, f'Could not save event: {e}')
            return render(request, 'create-event.html', {
                'categories': categories, 'venues': venues,
                'unread_notification_count': _unread_notif_count(request),
            })
    return render(request, 'create-event.html', {
        'categories': categories, 'venues': venues,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def event_list(request):
    events = Event.objects.select_related('category').all().order_by('-id')
    return render(request, 'event-list.html', {
        'events': events,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def event_detail(request, pk):
    event = get_object_or_404(Event, pk=pk)
    members = event.members.all()
    history = event.history.all()[:10]
    vendors = event.vendors.all()
    return render(request, 'event-detail.html', {
        'event': event, 'members': members, 'history': history, 'vendors': vendors,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def event_history_view(request, pk):
    event = get_object_or_404(Event, pk=pk)
    history = event.history.all()
    return render(request, 'event-history.html', {
        'event': event, 'history': history,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    categories = EventCategory.objects.all()
    venues = Venue.objects.filter(status='available')
    if request.method == 'POST':
        category_id = request.POST.get('category')
        if not category_id:
            messages.error(request, 'Please select a category.')
            return render(request, 'edit-event.html', {
                'event': event, 'categories': categories, 'venues': venues,
                'unread_notification_count': _unread_notif_count(request),
            })
        try:
            venue_ref_id = request.POST.get('venue_ref') or None
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            start_time = request.POST.get('start_time') or None
            end_time = request.POST.get('end_time') or None
            max_attendees = request.POST.get('max_attendees') or 0

            venue_ref_obj = Venue.objects.filter(pk=venue_ref_id).first() if venue_ref_id else None

            if venue_ref_obj:
                capacity_warning = _check_venue_capacity(venue_ref_obj, max_attendees)
                if capacity_warning:
                    messages.error(request, capacity_warning)
                    return render(request, 'edit-event.html', {
                        'event': event, 'categories': categories, 'venues': venues,
                        'unread_notification_count': _unread_notif_count(request),
                    })

                conflict_warning = _check_venue_schedule_conflict(
                    venue_ref_obj, start_date, end_date, start_time, end_time, exclude_pk=event.pk
                )
                if conflict_warning:
                    messages.error(request, conflict_warning)
                    return render(request, 'edit-event.html', {
                        'event': event, 'categories': categories, 'venues': venues,
                        'unread_notification_count': _unread_notif_count(request),
                    })

            old_status = event.status
            event.name = request.POST.get('name')
            event.category_id = category_id
            event.description = request.POST.get('description')
            event.priority = request.POST.get('priority') or 1
            event.scheduled_status = request.POST.get('scheduled_status')
            event.venue = request.POST.get('venue')
            event.venue_ref_id = venue_ref_id
            event.start_date = start_date
            event.end_date = end_date
            event.location = request.POST.get('location')
            event.points = request.POST.get('points') or 0
            event.max_attendees = max_attendees
            event.registration_deadline = request.POST.get('registration_deadline') or None
            event.status = request.POST.get('status')
            if request.FILES.get('image'):
                event.image = request.FILES.get('image')
            event.session_name = request.POST.get('session_name')
            event.speaker_name = request.POST.get('speaker_name')
            event.start_time = start_time
            event.end_time = end_time
            event.venue_name = request.POST.get('venue_name')
            event.budget = request.POST.get('budget') or 0
            event.sponsors = request.POST.get('sponsors', '')
            event.save()

            action = 'status_changed' if old_status != event.status else 'updated'
            EventHistory.objects.create(
                event=event, action=action, changed_by=request.user,
                notes=f'Updated. Status: {old_status} -> {event.status}' if action == 'status_changed' else 'Event details updated.'
            )

            if action == 'status_changed':
                for member in event.members.filter(user__isnull=False):
                    notify(member.user, f'Event "{event.name}" status changed to {event.status}.', link='/participant-dashboard/')

            messages.success(request, 'Event updated successfully!')
            return redirect('event_list')
        except Exception as e:
            messages.error(request, f'Could not update event: {e}')
            return render(request, 'edit-event.html', {
                'event': event, 'categories': categories, 'venues': venues,
                'unread_notification_count': _unread_notif_count(request),
            })
    return render(request, 'edit-event.html', {
        'event': event, 'categories': categories, 'venues': venues,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def delete_event_confirm(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        event.delete()
        messages.success(request, 'Event deleted.')
        return redirect('event_list')
    return render(request, 'delete-event.html', {
        'event': event,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def update_event_status(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        old_status = event.status
        new_status = request.POST.get('status')
        event.status = new_status
        event.save()
        EventHistory.objects.create(
            event=event, action='status_changed', changed_by=request.user,
            notes=f'Status changed: {old_status} -> {new_status}'
        )
        for member in event.members.filter(user__isnull=False):
            notify(member.user, f'Event "{event.name}" status changed to {new_status}.', link='/participant-dashboard/')
        messages.success(request, f'Event status updated to {new_status}.')
        return redirect('event_list')
    return render(request, 'update-event-status.html', {
        'event': event,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Event Members ----------

@admin_required
def add_event_member(request):
    events = Event.objects.all()
    if request.method == 'POST':
        event_id = request.POST.get('event')
        if not event_id:
            messages.error(request, 'Please select an event.')
            return render(request, 'add-event-member.html', {
                'events': events,
                'unread_notification_count': _unread_notif_count(request),
            })

        role_value = request.POST.get('role')

        member = EventMember.objects.create(
            event_id=event_id,
            name=request.POST.get('name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            role=role_value,
            department=request.POST.get('department', ''),
            academic_year=request.POST.get('academic_year', ''),
            registration_status=request.POST.get('registration_status') or 'confirmed',
        )

        if member.requires_qr:
            generate_qr_for_member(member)
            messages.success(request, 'Member added successfully! QR ticket generated.')
        else:
            messages.success(request, 'Member added successfully! (No QR ticket needed for this role.)')

        return redirect('add_event_member')
    return render(request, 'add-event-member.html', {
        'events': events,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def remove_event_member(request, pk):
    member = get_object_or_404(EventMember, pk=pk)
    if request.method == 'POST':
        event_pk = member.event.pk
        member.delete()
        messages.success(request, 'Member removed.')
        return redirect('event_detail', pk=event_pk)
    return render(request, 'remove-event-member.html', {
        'member': member,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def update_member_registration_status(request, pk):
    member = get_object_or_404(EventMember, pk=pk)
    if request.method == 'POST':
        member.registration_status = request.POST.get('registration_status')
        member.save()
        messages.success(request, f'Registration status for {member.name} updated.')
        return redirect('event_detail', pk=member.event.pk)
    return render(request, 'update-member-status.html', {
        'member': member,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def joined_events(request):
    events = Event.objects.filter(status__in=['active', 'upcoming'])
    return render(request, 'joined-events.html', {
        'events': events,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def completed_events(request):
    events = Event.objects.filter(status='completed')
    return render(request, 'completed-events.html', {
        'events': events,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def complete_event_user_list(request, pk):
    event = get_object_or_404(Event, pk=pk)
    members = event.members.all()
    return render(request, 'complete-event-user-list.html', {
        'event': event, 'members': members,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Attendance (User Marks) ----------

@admin_required
def create_user_mark(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)

    if event.status == 'upcoming':
        messages.error(request, 'Attendance cannot be marked yet — this event has not started. Attendance becomes available once the event is Active or Completed.')
        return redirect('event_detail', pk=event.pk)

    members = event.members.filter(registration_status='confirmed')

    member_status = {}
    for m in members:
        latest = m.marks.order_by('-marked_at').first()
        member_status[m.id] = latest.status if latest else 'not_marked'

    if request.method == 'POST':
        awarded_this_save = 0
        for member in members:
            status = request.POST.get(f'status_{member.id}')
            if status not in ('present', 'absent'):
                continue

            latest = member.marks.order_by('-marked_at').first()
            if latest:
                latest.status = status
                latest.save()
            else:
                UserMark.objects.create(member=member, status=status)

            if status == 'present' and not member.points_awarded:
                if member.user:
                    profile, _ = Profile.objects.get_or_create(
                        user=member.user, defaults={'role': 'participant'}
                    )
                    profile.total_points += event.points
                    profile.save()
                member.points_awarded = True
                member.save()
                awarded_this_save += 1

        if awarded_this_save:
            messages.success(request, f'Attendance saved. {awarded_this_save} participant(s) awarded {event.points} points each.')
        else:
            messages.success(request, 'Attendance saved successfully.')
        return redirect('user_mark_list', event_pk=event.pk)

    return render(request, 'create-user-mark.html', {
        'event': event,
        'members': members,
        'member_status': member_status,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def user_mark_list(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    members = event.members.filter(registration_status='confirmed')

    total_registered = members.count()
    present_count = 0
    absent_count = 0
    rows = []

    for m in members:
        latest = m.marks.order_by('-marked_at').first()
        status = latest.status if latest else 'not_marked'
        if status == 'present':
            present_count += 1
        elif status == 'absent':
            absent_count += 1
        rows.append({
            'member': m,
            'status': status,
            'marked_at': latest.marked_at if latest else None,
        })

    not_marked_count = total_registered - present_count - absent_count
    attendance_pct = round((present_count / total_registered) * 100, 1) if total_registered else 0

    return render(request, 'user-mark-list.html', {
        'event': event,
        'rows': rows,
        'total_registered': total_registered,
        'present_count': present_count,
        'absent_count': absent_count,
        'not_marked_count': not_marked_count,
        'attendance_pct': attendance_pct,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def absent_user_list(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    marks = UserMark.objects.filter(member__event=event, status='absent').select_related('member')
    return render(request, 'absent-user-list.html', {
        'event': event, 'marks': marks,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def contact(request):
    profile = getattr(request.user, 'profile', None)
    is_admin = profile and profile.role == 'admin'
    template_name = 'admin-contact.html' if is_admin else 'contact.html'
    home_url = 'dashboard' if is_admin else 'participant_dashboard'

    if request.method == 'POST':
        name = request.POST.get('name') or request.user.get_full_name() or request.user.username
        email = request.POST.get('email') or request.user.email
        subject = request.POST.get('subject', '')
        message_text = request.POST.get('message', '')

        if not message_text.strip():
            messages.error(request, 'Please enter a message before submitting.')
            return render(request, template_name, {
                'home_url': home_url,
                'unread_notification_count': _unread_notif_count(request),
            })

        inquiry = ContactMessage.objects.create(
            name=name,
            email=email,
            subject=subject,
            message=message_text,
            submitted_by=request.user,
        )
        notify_admins(
            f'{name}: {subject or message_text[:40]}',
            link=f'/contact-messages/{inquiry.id}/'
        )

        messages.success(request, 'Message sent successfully!')
        return redirect('contact')

    return render(request, template_name, {
        'home_url': home_url,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def contact_message_list(request):
    inquiries = ContactMessage.objects.select_related('submitted_by').all()
    return render(request, 'contact-message-list.html', {
        'inquiries': inquiries,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def contact_message_detail(request, pk):
    inquiry = get_object_or_404(ContactMessage, pk=pk)

    if inquiry.status == 'new':
        inquiry.status = 'read'
        inquiry.save()

    Notification.objects.filter(
        user=request.user,
        link=f'/contact-messages/{inquiry.id}/',
        is_read=False
    ).update(is_read=True)

    if request.method == 'POST':
        reply_text = request.POST.get('reply', '').strip()
        if reply_text:
            ContactMessageReply.objects.create(
                contact_message=inquiry,
                sender=request.user,
                is_admin_reply=True,
                message=reply_text,
            )
            if inquiry.submitted_by:
                notify(
                    inquiry.submitted_by,
                    f'Admin replied: {inquiry.subject or "your message"}',
                    link=f'/my-inquiries/{inquiry.id}/'
                )
            messages.success(request, 'Reply sent.')
        return redirect('contact_message_detail', pk=inquiry.pk)

    replies = inquiry.replies.select_related('sender').all()
    return render(request, 'contact-message-detail.html', {
        'inquiry': inquiry,
        'replies': replies,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def mark_contact_reviewed(request, pk):
    inquiry = get_object_or_404(ContactMessage, pk=pk)
    inquiry.status = 'resolved'
    inquiry.save()
    Notification.objects.filter(
        link=f'/contact-messages/{inquiry.id}/',
        is_read=False
    ).update(is_read=True)
    messages.success(request, 'Marked as resolved.')
    return redirect('contact_message_detail', pk=inquiry.pk)


@login_required(login_url='/login/')
def participant_inquiry_list(request):
    inquiries = ContactMessage.objects.filter(submitted_by=request.user)
    return render(request, 'participant-inquiry-list.html', {
        'inquiries': inquiries,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def participant_inquiry_detail(request, pk):
    inquiry = get_object_or_404(ContactMessage, pk=pk, submitted_by=request.user)

    Notification.objects.filter(
        user=request.user,
        link=f'/my-inquiries/{inquiry.id}/',
        is_read=False
    ).update(is_read=True)

    if request.method == 'POST':
        reply_text = request.POST.get('reply', '').strip()
        if reply_text:
            ContactMessageReply.objects.create(
                contact_message=inquiry,
                sender=request.user,
                is_admin_reply=False,
                message=reply_text,
            )
            if inquiry.status == 'resolved':
                inquiry.status = 'read'
                inquiry.save()
            notify_admins(
                f'{request.user.username} replied: {inquiry.subject or "message"}',
                link=f'/contact-messages/{inquiry.id}/'
            )
            messages.success(request, 'Message sent.')
        return redirect('participant_inquiry_detail', pk=inquiry.pk)

    replies = inquiry.replies.select_related('sender').all()
    return render(request, 'participant-inquiry-detail.html', {
        'inquiry': inquiry,
        'replies': replies,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Event Wish ----------

@login_required(login_url='/login/')
def add_event_wish(request):
    if request.method == 'POST':
        wish = EventUserWish.objects.create(
            event_id=request.POST.get('event'),
            user_id=request.POST.get('user'),
            status=request.POST.get('status'),
        )
        notify_admins(f'New event wish from {wish.user.username} for "{wish.event.name}".', link='/event-wish-list/')
        return redirect('event_wish_list')
    events = Event.objects.all()
    users = User.objects.all()
    return render(request, 'add-event-wish.html', {
        'events': events, 'users': users,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def event_wish_list(request):
    wishes = EventUserWish.objects.select_related('event', 'user').all()
    return render(request, 'event-wish-list.html', {
        'wishes': wishes,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def approve_wish(request, pk):
    wish = get_object_or_404(EventUserWish, pk=pk)
    wish.status = 'approved'
    wish.save()
    notify(wish.user, f'Your wish for "{wish.event.name}" was approved!', link='/participant-dashboard/')
    messages.success(request, 'Wish approved.')
    return redirect('event_wish_list')


@admin_required
def reject_wish(request, pk):
    wish = get_object_or_404(EventUserWish, pk=pk)
    wish.status = 'rejected'
    wish.save()
    notify(wish.user, f'Your wish for "{wish.event.name}" was rejected.', link='/participant-dashboard/')
    messages.success(request, 'Wish rejected.')
    return redirect('event_wish_list')


@admin_required
def remove_event_wish(request, pk):
    wish = get_object_or_404(EventUserWish, pk=pk)
    if request.method == 'POST':
        wish.delete()
        messages.success(request, 'Wish removed.')
        return redirect('event_wish_list')
    return render(request, 'remove-event-wish.html', {
        'wish': wish,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Vendors ----------

@admin_required
def vendor_list(request):
    vendors = Vendor.objects.all().order_by('-created_at')
    return render(request, 'vendor-list.html', {
        'vendors': vendors,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_vendor(request):
    events = Event.objects.all()
    if request.method == 'POST':
        Vendor.objects.create(
            name=request.POST.get('name'),
            contact_person=request.POST.get('contact_person'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            service_type=request.POST.get('service_type'),
            contract_status=request.POST.get('contract_status'),
            contract_document=request.FILES.get('contract_document'),
            contract_start_date=request.POST.get('contract_start_date') or None,
            contract_end_date=request.POST.get('contract_end_date') or None,
            event_id=request.POST.get('event') or None,
            notes=request.POST.get('notes'),
        )
        messages.success(request, 'Vendor onboarded successfully!')
        return redirect('vendor_list')
    return render(request, 'vendor-form.html', {
        'events': events, 'vendor': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_vendor(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    events = Event.objects.all()
    if request.method == 'POST':
        vendor.name = request.POST.get('name')
        vendor.contact_person = request.POST.get('contact_person')
        vendor.email = request.POST.get('email')
        vendor.phone = request.POST.get('phone')
        vendor.service_type = request.POST.get('service_type')
        vendor.contract_status = request.POST.get('contract_status')
        if request.FILES.get('contract_document'):
            vendor.contract_document = request.FILES.get('contract_document')
        vendor.contract_start_date = request.POST.get('contract_start_date') or None
        vendor.contract_end_date = request.POST.get('contract_end_date') or None
        vendor.event_id = request.POST.get('event') or None
        vendor.notes = request.POST.get('notes')
        vendor.save()
        messages.success(request, 'Vendor updated successfully!')
        return redirect('vendor_list')
    return render(request, 'vendor-form.html', {
        'events': events, 'vendor': vendor,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def delete_vendor_confirm(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        vendor.delete()
        messages.success(request, 'Vendor removed.')
        return redirect('vendor_list')
    return render(request, 'vendor-delete.html', {
        'vendor': vendor,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Budget & Financial Tracking ----------

@admin_required
def budget_overview(request):
    events = Event.objects.select_related('category').all().order_by('-budget')
    total_budget = events.aggregate(total=Sum('budget'))['total'] or 0

    total_projected = Expense.objects.aggregate(total=Sum('projected_amount'))['total'] or 0
    total_actual = Expense.objects.filter(status='approved').aggregate(total=Sum('actual_amount'))['total'] or 0
    total_sponsorship = SponsorshipRevenue.objects.aggregate(total=Sum('amount'))['total'] or 0
    total_received = SponsorshipRevenue.objects.filter(status='received').aggregate(total=Sum('amount'))['total'] or 0

    utilization_pct = 0
    if total_budget > 0:
        utilization_pct = round((total_actual / total_budget) * 100, 1)

    category_breakdown = (
        Expense.objects.values('category')
        .annotate(projected=Sum('projected_amount'), actual=Sum('actual_amount'))
        .order_by('-projected')
    )

    pending_approvals = Expense.objects.filter(status='pending').select_related('event', 'requested_by')

    events_with_sponsors = events.exclude(sponsors='').exclude(sponsors__isnull=True)

    return render(request, 'budget-overview.html', {
        'events': events,
        'total_budget': total_budget,
        'total_projected': total_projected,
        'total_actual': total_actual,
        'total_sponsorship': total_sponsorship,
        'total_received': total_received,
        'utilization_pct': utilization_pct,
        'category_breakdown': category_breakdown,
        'pending_approvals': pending_approvals,
        'events_with_sponsors': events_with_sponsors,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_expense(request):
    events = Event.objects.all()
    if request.method == 'POST':
        event_id = request.POST.get('event')
        if not event_id:
            messages.error(request, 'Please select an event.')
            return render(request, 'expense-form.html', {
                'events': events, 'expense': None,
                'unread_notification_count': _unread_notif_count(request),
            })

        try:
            projected = float(request.POST.get('projected_amount') or 0)
        except ValueError:
            projected = 0

        try:
            actual = float(request.POST.get('actual_amount') or 0)
        except ValueError:
            actual = 0

        status = 'pending' if projected >= Expense.HIGH_VALUE_THRESHOLD else 'approved'

        expense = Expense.objects.create(
            event_id=event_id,
            category=request.POST.get('category'),
            description=request.POST.get('description'),
            projected_amount=projected,
            actual_amount=actual,
            invoice=request.FILES.get('invoice'),
            status=status,
            requested_by=request.user,
            approved_by=request.user if status == 'approved' else None,
        )

        if status == 'pending':
            messages.warning(request, f'This expense (₹{expense.projected_amount}) exceeds ₹{Expense.HIGH_VALUE_THRESHOLD} and requires approval before it counts toward actual spend.')
        else:
            messages.success(request, 'Expense recorded successfully!')

        return redirect('budget_overview')

    return render(request, 'expense-form.html', {
        'events': events, 'expense': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def approve_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    expense.status = 'approved'
    expense.approved_by = request.user
    expense.save()
    messages.success(request, f'Expense of ₹{expense.projected_amount} approved.')
    return redirect('budget_overview')


@admin_required
def reject_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    expense.status = 'rejected'
    expense.approved_by = request.user
    expense.save()
    messages.success(request, 'Expense rejected.')
    return redirect('budget_overview')


@admin_required
def create_sponsorship(request):
    events = Event.objects.all()
    sponsors = Sponsor.objects.all()
    if request.method == 'POST':
        event_id = request.POST.get('event')
        if not event_id:
            messages.error(request, 'Please select an event.')
            return render(request, 'sponsorship-form.html', {
                'events': events, 'sponsors': sponsors,
                'unread_notification_count': _unread_notif_count(request),
            })

        sponsor_ref_id = request.POST.get('sponsor_ref') or None
        sponsor_name = request.POST.get('sponsor_name', '')

        if sponsor_ref_id and not sponsor_name:
            sponsor_ref = get_object_or_404(Sponsor, pk=sponsor_ref_id)
            sponsor_name = sponsor_ref.name

        SponsorshipRevenue.objects.create(
            event_id=event_id,
            sponsor_ref_id=sponsor_ref_id,
            sponsor_name=sponsor_name,
            amount=request.POST.get('amount') or 0,
            status=request.POST.get('status'),
        )
        messages.success(request, 'Sponsorship revenue recorded!')
        return redirect('budget_overview')

    return render(request, 'sponsorship-form.html', {
        'events': events, 'sponsors': sponsors,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Venue Management ----------

@admin_required
def venue_list(request):
    venues = Venue.objects.all()
    return render(request, 'venue-list.html', {
        'venues': venues,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_venue(request):
    if request.method == 'POST':
        Venue.objects.create(
            name=request.POST.get('name'),
            location=request.POST.get('location'),
            capacity=request.POST.get('capacity') or 0,
            venue_type=request.POST.get('venue_type') or 'other',
            status=request.POST.get('status') or 'available',
            description=request.POST.get('description', ''),
            has_projector='has_projector' in request.POST,
            has_ac='has_ac' in request.POST,
            has_sound_system='has_sound_system' in request.POST,
            has_wifi='has_wifi' in request.POST,
            has_stage='has_stage' in request.POST,
        )
        messages.success(request, 'Venue registered successfully!')
        return redirect('venue_list')
    return render(request, 'venue-form.html', {
        'venue': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_venue(request, pk):
    venue = get_object_or_404(Venue, pk=pk)
    if request.method == 'POST':
        venue.name = request.POST.get('name')
        venue.location = request.POST.get('location')
        venue.capacity = request.POST.get('capacity') or 0
        venue.venue_type = request.POST.get('venue_type') or 'other'
        venue.status = request.POST.get('status') or 'available'
        venue.description = request.POST.get('description', '')
        venue.has_projector = 'has_projector' in request.POST
        venue.has_ac = 'has_ac' in request.POST
        venue.has_sound_system = 'has_sound_system' in request.POST
        venue.has_wifi = 'has_wifi' in request.POST
        venue.has_stage = 'has_stage' in request.POST
        venue.save()
        messages.success(request, 'Venue updated successfully!')
        return redirect('venue_list')
    return render(request, 'venue-form.html', {
        'venue': venue,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def delete_venue_confirm(request, pk):
    venue = get_object_or_404(Venue, pk=pk)
    if request.method == 'POST':
        venue.delete()
        messages.success(request, 'Venue removed.')
        return redirect('venue_list')
    return render(request, 'venue-delete.html', {
        'venue': venue,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Resource Allocation ----------

@admin_required
def resource_list(request):
    resources = Resource.objects.all()
    allocations = ResourceAllocation.objects.select_related('resource', 'event').order_by('-allocated_at')[:50]
    return render(request, 'resource-list.html', {
        'resources': resources,
        'allocations': allocations,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_resource(request):
    if request.method == 'POST':
        Resource.objects.create(
            name=request.POST.get('name'),
            category=request.POST.get('category') or 'other',
            total_quantity=request.POST.get('total_quantity') or 1,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, 'Resource added to inventory!')
        return redirect('resource_list')
    return render(request, 'resource-form.html', {
        'resource': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_resource(request, pk):
    resource = get_object_or_404(Resource, pk=pk)
    if request.method == 'POST':
        resource.name = request.POST.get('name')
        resource.category = request.POST.get('category') or 'other'
        resource.total_quantity = request.POST.get('total_quantity') or 1
        resource.notes = request.POST.get('notes', '')
        resource.save()
        messages.success(request, 'Resource updated successfully!')
        return redirect('resource_list')
    return render(request, 'resource-form.html', {
        'resource': resource,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def delete_resource_confirm(request, pk):
    resource = get_object_or_404(Resource, pk=pk)
    if request.method == 'POST':
        resource.delete()
        messages.success(request, 'Resource removed from inventory.')
        return redirect('resource_list')
    return render(request, 'resource-delete.html', {
        'resource': resource,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def allocate_resource(request):
    resources = Resource.objects.all()
    events = Event.objects.filter(status__in=['upcoming', 'active'])
    if request.method == 'POST':
        resource_id = request.POST.get('resource')
        event_id = request.POST.get('event')
        try:
            quantity = int(request.POST.get('quantity') or 1)
        except ValueError:
            quantity = 1

        if not resource_id or not event_id:
            messages.error(request, 'Please select both a resource and an event.')
            return render(request, 'resource-allocate.html', {
                'resources': resources, 'events': events,
                'unread_notification_count': _unread_notif_count(request),
            })

        resource = get_object_or_404(Resource, pk=resource_id)
        if quantity > resource.available_quantity:
            messages.error(request, f'Only {resource.available_quantity} of "{resource.name}" available — cannot allocate {quantity}.')
            return render(request, 'resource-allocate.html', {
                'resources': resources, 'events': events,
                'unread_notification_count': _unread_notif_count(request),
            })

        ResourceAllocation.objects.create(resource_id=resource_id, event_id=event_id, quantity=quantity)
        messages.success(request, f'{quantity}x "{resource.name}" allocated successfully!')
        return redirect('resource_list')

    return render(request, 'resource-allocate.html', {
        'resources': resources, 'events': events,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def remove_resource_allocation(request, pk):
    allocation = get_object_or_404(ResourceAllocation, pk=pk)
    if request.method == 'POST':
        allocation.delete()
        messages.success(request, 'Allocation removed.')
        return redirect('resource_list')
    return render(request, 'resource-allocation-delete.html', {
        'allocation': allocation,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Sponsor Management ----------

@admin_required
def sponsor_list(request):
    sponsors = Sponsor.objects.all()
    return render(request, 'sponsor-list.html', {
        'sponsors': sponsors,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_sponsor(request):
    if request.method == 'POST':
        Sponsor.objects.create(
            name=request.POST.get('name'),
            logo=request.FILES.get('logo'),
            contact_person=request.POST.get('contact_person', ''),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            website=request.POST.get('website', ''),
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, 'Sponsor registered successfully!')
        return redirect('sponsor_list')
    return render(request, 'sponsor-form.html', {
        'sponsor': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_sponsor(request, pk):
    sponsor = get_object_or_404(Sponsor, pk=pk)
    if request.method == 'POST':
        sponsor.name = request.POST.get('name')
        if request.FILES.get('logo'):
            sponsor.logo = request.FILES.get('logo')
        sponsor.contact_person = request.POST.get('contact_person', '')
        sponsor.email = request.POST.get('email', '')
        sponsor.phone = request.POST.get('phone', '')
        sponsor.website = request.POST.get('website', '')
        sponsor.notes = request.POST.get('notes', '')
        sponsor.save()
        messages.success(request, 'Sponsor updated successfully!')
        return redirect('sponsor_list')
    return render(request, 'sponsor-form.html', {
        'sponsor': sponsor,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def delete_sponsor_confirm(request, pk):
    sponsor = get_object_or_404(Sponsor, pk=pk)
    if request.method == 'POST':
        sponsor.delete()
        messages.success(request, 'Sponsor removed.')
        return redirect('sponsor_list')
    return render(request, 'sponsor-delete.html', {
        'sponsor': sponsor,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- Reports & Export ----------

@admin_required
def reports_home(request):
    return render(request, 'reports.html', {
        'unread_notification_count': _unread_notif_count(request),
    })


def _pdf_table_response(filename, title, headers, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                             leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 10)]

    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dcdfe6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7fd')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


def _xlsx_response(filename, sheet_title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _events_rows():
    headers = ['ID', 'Name', 'Category', 'Status', 'Start Date', 'End Date', 'Venue', 'Max Attendees', 'Registered', 'Budget']
    rows = []
    for ev in Event.objects.select_related('category').all():
        rows.append([
            ev.id, ev.name, ev.category.name, ev.status,
            str(ev.start_date), str(ev.end_date), ev.venue,
            ev.max_attendees, ev.members.count(), str(ev.budget),
        ])
    return headers, rows


def _members_rows():
    headers = ['ID', 'Name', 'Email', 'Phone', 'Event', 'Role', 'Department', 'Academic Year', 'Ticket Code', 'Registration Status', 'Joined At']
    rows = []
    for m in EventMember.objects.select_related('event').all():
        rows.append([
            m.id, m.name, m.email, m.phone, m.event.name,
            m.get_role_display(), m.department, m.academic_year, m.ticket_code or '-',
            m.get_registration_status_display(), str(m.joined_at),
        ])
    return headers, rows


def _contacts_rows():
    headers = ['ID', 'Name', 'Email', 'Subject', 'Message', 'Status', 'Submitted At']
    rows = []
    for c in ContactMessage.objects.all():
        rows.append([
            c.id, c.name, c.email, c.subject, c.message,
            c.get_status_display(), str(c.submitted_at),
        ])
    return headers, rows


@admin_required
def export_events_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="events_report.csv"'
    writer = csv.writer(response)
    headers, rows = _events_rows()
    writer.writerow(headers)
    writer.writerows(rows)
    return response


@admin_required
def export_events_xlsx(request):
    headers, rows = _events_rows()
    return _xlsx_response('events_report.xlsx', 'Events', headers, rows)


@admin_required
def export_events_pdf(request):
    headers, rows = _events_rows()
    return _pdf_table_response('events_report.pdf', 'Event Report', headers, rows)


@admin_required
def export_members_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="members_report.csv"'
    writer = csv.writer(response)
    headers, rows = _members_rows()
    writer.writerow(headers)
    writer.writerows(rows)
    return response


@admin_required
def export_members_xlsx(request):
    headers, rows = _members_rows()
    return _xlsx_response('members_report.xlsx', 'Members', headers, rows)


@admin_required
def export_members_pdf(request):
    headers, rows = _members_rows()
    return _pdf_table_response('members_report.pdf', 'Member Report', headers, rows)


@admin_required
def export_contacts_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="contact_messages_report.csv"'
    writer = csv.writer(response)
    headers, rows = _contacts_rows()
    writer.writerow(headers)
    writer.writerows(rows)
    return response


@admin_required
def export_contacts_xlsx(request):
    headers, rows = _contacts_rows()
    return _xlsx_response('contact_messages_report.xlsx', 'Contacts', headers, rows)


@admin_required
def export_contacts_pdf(request):
    headers, rows = _contacts_rows()
    return _pdf_table_response('contact_messages_report.pdf', 'Contact Messages Report', headers, rows)


# ---------- Policies & Guidelines ----------

@admin_required
def policy_list(request):
    policies = Policy.objects.all()
    return render(request, 'policy-list.html', {
        'policies': policies,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def create_policy(request):
    if request.method == 'POST':
        Policy.objects.create(
            title=request.POST.get('title'),
            category=request.POST.get('category') or 'other',
            content=request.POST.get('content'),
            status=request.POST.get('status') or 'active',
            created_by=request.user,
        )
        messages.success(request, 'Policy created successfully!')
        return redirect('policy_list')
    return render(request, 'policy-form.html', {
        'policy': None,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def edit_policy(request, pk):
    policy = get_object_or_404(Policy, pk=pk)
    if request.method == 'POST':
        policy.title = request.POST.get('title')
        policy.category = request.POST.get('category') or 'other'
        policy.content = request.POST.get('content')
        policy.status = request.POST.get('status') or 'active'
        policy.save()
        messages.success(request, 'Policy updated successfully!')
        return redirect('policy_list')
    return render(request, 'policy-form.html', {
        'policy': policy,
        'unread_notification_count': _unread_notif_count(request),
    })


@admin_required
def toggle_policy_status(request, pk):
    policy = get_object_or_404(Policy, pk=pk)
    policy.status = 'inactive' if policy.status == 'active' else 'active'
    policy.save()
    messages.success(request, f'"{policy.title}" is now {policy.get_status_display()}.')
    return redirect('policy_list')


@admin_required
def delete_policy_confirm(request, pk):
    policy = get_object_or_404(Policy, pk=pk)
    if request.method == 'POST':
        policy.delete()
        messages.success(request, 'Policy deleted.')
        return redirect('policy_list')
    return render(request, 'policy-delete.html', {
        'policy': policy,
        'unread_notification_count': _unread_notif_count(request),
    })


@login_required(login_url='/login/')
def participant_policy_list(request):
    policies = Policy.objects.filter(status='active')
    return render(request, 'participant-policy-list.html', {
        'policies': policies,
        'unread_notification_count': _unread_notif_count(request),
    })


# ---------- FAQ Chatbot ----------

FAQ_CONTEXT = """
You are Chatty, the AI assistant for EventSphere — a college Event Registration and Management Platform.

YOUR SCOPE:
You ONLY answer questions related to EventSphere — its features, how to use it, its data (events, venues, vendors, resources, sponsors, registrations, tickets), and reasonable questions about how the project could be extended in the future.

If a question is unrelated to EventSphere (general knowledge, other topics, unrelated coding questions, homework help, etc.), reply briefly and politely that you are specific to the EventSphere project and can only help with questions related to it. Do not attempt to answer unrelated questions, even partially.

FORMATTING RULES:
- Use numbered steps or bullet points whenever explaining a process or listing multiple items.
- Keep answers concise and well-organized. Avoid long unbroken paragraphs.
- Bold key terms sparingly if it helps readability.

HOW TO HANDLE DIFFERENT QUESTION TYPES:

1. Questions about existing features (how do I register, where's my ticket, how does QR check-in work):
Answer using the feature descriptions below.

2. Questions about live data (what events exist, how many resources are available, which venues exist):
Answer STRICTLY using the LIVE DATA section below, which reflects the actual current database contents at this exact moment. This data is always accurate and up-to-date — if something is not listed there, it does not currently exist, has been removed, or was deleted by an administrator. Never mention, describe, or assume the existence of any event, venue, resource, vendor, category, or sponsor that is not explicitly present in the LIVE DATA section, even if it was discussed earlier in this conversation or seems familiar. If the live data shows nothing available in a category, state clearly that nothing is currently available — do not guess, invent, or recall from memory.

3. Questions about features that DO NOT currently exist in this project (e.g. "how would OTP-based registration work here?", "can you add WhatsApp login?"):
Be honest that the feature is not currently implemented, then answer helpfully as a knowledgeable assistant discussing how it COULD be added to this project — giving a clear, step-by-step technical explanation. Never claim a feature exists if it doesn't.

EXISTING FEATURES:

1. How do I register for an event?
Go to "Events Schedule" in the sidebar, click on any event to view details, then click "Confirm Registration." Registration happens instantly and a QR ticket is generated automatically.

2. Where can I find my QR ticket?
Go to the Dashboard and check "My Registrations & Tickets" (or "My Tickets" in the sidebar). Each registered event shows a QR code and a unique ticket code.

3. Can I register for an event that already started?
No — if the registration deadline has passed, or the event is marked Completed, registration is blocked.

4. What happens if an event is full?
If max attendee capacity is reached, registration is blocked with a clear message.

5. How do I cancel my registration?
Cancellation is available through the participant dashboard, if enabled for that event.

6. How does QR check-in work?
Admin selects the event, then manually enters (or scans) the participant's unique ticket code (e.g. MEMBER-7X92K4) to verify entry. Entry and Exit modes are both supported, and duplicate check-ins are automatically blocked.

7. What if my ticket code doesn't match?
The system checks that the ticket belongs to the currently selected event before allowing check-in.

8. Can I register for multiple events?
Yes, as long as each event is still open for registration and has available capacity.

9. How do I update my profile?
Go to "My Profile" in the sidebar to update your name, email, phone number, and profile photo.

10. Who do I contact for issues?
Use the "Contact" page in the sidebar to reach the event administrators.

11. What roles exist on the platform?
Admin (manages events, vendors, venues, resources, sponsors, budgets) and Participant (browses and registers for events).
"""


def _build_live_data_context():
    today = timezone.now().date()

    upcoming_events = (
        Event.objects.filter(end_date__gte=today)
        .exclude(status='completed')
        .select_related('category')
        .order_by('start_date')[:15]
    )
    event_lines = [
        f"- {e.name} ({e.category.name}) — {e.start_date} to {e.end_date}, status: {e.status}, venue: {e.venue or 'TBA'}"
        for e in upcoming_events
    ]

    venues = Venue.objects.filter(status='available')[:15]
    venue_lines = [
        f"- {v.name} ({v.get_venue_type_display()}), capacity: {v.capacity}, location: {v.location or 'N/A'}"
        for v in venues
    ]

    resources = Resource.objects.all()[:15]
    resource_lines = [
        f"- {r.name} ({r.get_category_display()}): {r.available_quantity} available out of {r.total_quantity}"
        for r in resources
    ]

    vendors = Vendor.objects.all()[:15]
    vendor_lines = [
        f"- {v.name} ({v.get_service_type_display()}), contract status: {v.get_contract_status_display()}"
        for v in vendors
    ]

    sponsors = Sponsor.objects.all()[:15]
    sponsor_lines = [f"- {s.name}" for s in sponsors]

    categories = EventCategory.objects.filter(status='active')[:15]
    category_lines = [f"- {c.name}" for c in categories]

    return f"""
LIVE DATA (current as of this exact moment — use this to answer "what's available" type questions):

Upcoming/Active Events:
{chr(10).join(event_lines) if event_lines else "No upcoming events currently."}

Available Venues:
{chr(10).join(venue_lines) if venue_lines else "No venues currently marked available."}

Resources & Availability:
{chr(10).join(resource_lines) if resource_lines else "No resources currently in inventory."}

Vendors:
{chr(10).join(vendor_lines) if vendor_lines else "No vendors currently onboarded."}

Sponsors:
{chr(10).join(sponsor_lines) if sponsor_lines else "No sponsors currently registered."}

Active Categories:
{chr(10).join(category_lines) if category_lines else "No active categories currently."}
"""


@login_required(login_url='/login/')
def chatbot_faq(request):
    if request.method == 'POST':
        user_question = request.POST.get('question', '').strip()

        if not user_question:
            return JsonResponse({'answer': "Please type a question."})

        if not settings.GROQ_API_KEY:
            return JsonResponse({'answer': "Chatbot is not configured yet. Please contact the admin."})

        try:
            live_data = _build_live_data_context()
            full_system_prompt = FAQ_CONTEXT + "\n\n" + live_data

            client = Groq(api_key=settings.GROQ_API_KEY)
            completion = client.chat.completions.create(
                model="openai/gpt-oss-20b",
                messages=[
                    {"role": "system", "content": full_system_prompt},
                    {"role": "user", "content": user_question},
                ],
                max_tokens=350,
                temperature=0.4,
            )
            answer = completion.choices[0].message.content
        except Exception:
            answer = "Sorry, I couldn't process that right now. Please try again shortly."

        return JsonResponse({'answer': answer})

    return render(request, 'chatbot-faq.html', {
        'unread_notification_count': _unread_notif_count(request),
    })